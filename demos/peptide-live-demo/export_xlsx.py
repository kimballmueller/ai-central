#!/usr/bin/env python3
"""
Export the live SQLite sweep to an Excel workbook in Brandon's v3 shape.

The point: the database is the source of truth and the sheet is a generated
artifact. He keeps the sheet he likes for analysis and sharing, and stops
maintaining it by hand.

Same conventions as scripts/peptide_migrate.py -- four METHOD types, formula
columns never written to, purple/white banding, green best-value cells, Full
Kits split at 10+ units, live A1 summary. Plus two things the workbook alone
could never do: a Price History tab and a Changes tab.

  python3 demos/peptide-live-demo/export_xlsx.py
"""

import argparse
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BAND_A, BAND_B, GREEN = "FFFFFFFF", "FFEDE9FE", "FF90EE90"
HEADERS = ["PEPTIDE", "VENDOR", "METHOD", "MG/VIAL", "MG/CAPSULE", "CAPS/BOTTLE",
           "MG/ML", "ML/BOTTLE", "# VIALS / BOTTLES", "MG TOTAL", "TOTAL $",
           "$/VIAL", "$/MG", "$/ML", "NOTES", "SCRAPED DATE"]
WIDTHS = [38, 26, 10, 10, 12, 12, 10, 11, 17, 11, 12, 11, 12, 11, 34, 14]
FORMULAS = {
    10: '=IF(OR($C{r}="",$I{r}=""),"",IF($C{r}="Oral",$I{r}*$F{r}*$E{r},'
        'IF(OR($C{r}="LiqDrop",$C{r}="Nasal"),$I{r}*$H{r}*$G{r},$I{r}*$D{r})))',
    12: '=IFERROR($K{r}/$I{r},"")',
    13: '=IFERROR($K{r}/$J{r},"")',
    14: '=IFERROR($K{r}/($I{r}*$H{r}),"")',
}


def sheet(wb, title, first=True):
    ws = wb.create_sheet(title) if not first else wb.active
    ws.title = title
    return ws


def write_tracker(ws, rows, summary):
    ws["A1"] = summary
    ws["A1"].font = Font(bold=True, size=11)
    for c, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(2, c, h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, r in enumerate(rows):
        rw = 3 + i
        band = BAND_A if i % 2 == 0 else BAND_B
        ws.cell(rw, 1, r["compound"])
        ws.cell(rw, 2, r["vendor"])
        ws.cell(rw, 3, r["method"])
        if r["method"] == "Inj":
            ws.cell(rw, 4, r["mg"])
        elif r["method"] == "LiqDrop":
            ws.cell(rw, 7, r["mg_ml"])
            ws.cell(rw, 8, r["ml"])
        ws.cell(rw, 9, r["units"])
        ws.cell(rw, 11, r["price"]).number_format = '"$"#,##0.00'
        if r["notes"]:
            ws.cell(rw, 15, r["notes"])
        ws.cell(rw, 16, r["scraped"]).number_format = "yyyy-mm-dd hh:mm"
        for col, tpl in FORMULAS.items():
            ws.cell(rw, col, tpl.format(r=rw))
        ws.cell(rw, 12).number_format = '"$"#,##0.00'
        ws.cell(rw, 13).number_format = '"$"#,##0.0000'
        ws.cell(rw, 14).number_format = '"$"#,##0.0000'
        for col in range(1, 17):
            ws.cell(rw, col).fill = PatternFill("solid", fgColor=band)

    # Best-value pass: cheapest $/mg per (compound, method, per-unit size).
    groups = defaultdict(list)
    for i, r in enumerate(rows):
        if r.get("per_mg"):
            groups[(r["compound"], r["method"], r["mg"])].append((i, r["per_mg"]))
    greens = 0
    for members in groups.values():
        lo = min(v for _, v in members)
        for i, v in members:
            if v == lo:
                ws.cell(3 + i, 13).fill = PatternFill("solid", fgColor=GREEN)
                greens += 1

    last = 2 + len(rows)
    ws.auto_filter.ref = f"A2:P{last}"
    ws.freeze_panes = "A3"
    return greens


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default="demos/peptide-live-demo/prices.db")
    ap.add_argument("--out", default="demos/peptide-live-demo/Peptide_Prices_LIVE.xlsx")
    args = ap.parse_args()

    db = sqlite3.connect(args.db)
    db.row_factory = sqlite3.Row
    latest = db.execute("SELECT MAX(id) FROM sweeps").fetchone()[0]
    swept = db.execute("SELECT started_at FROM sweeps WHERE id=?", (latest,)).fetchone()[0]
    scraped = datetime.fromisoformat(swept).replace(tzinfo=None)

    rows = []
    for r in db.execute("""
            SELECT s.compound, s.vendor, s.method, s.mg, s.units, s.product, s.variant,
                   p.price, p.regular, p.on_sale, p.in_stock
            FROM skus s JOIN price_snapshots p ON p.sku_id = s.id
            WHERE p.sweep_id = ? AND s.mg IS NOT NULL AND s.mg > 0
            ORDER BY s.compound, s.vendor, s.mg, s.units""", (latest,)):
        notes = []
        if r["on_sale"] and r["regular"]:
            notes.append(f"SALE, was ${r['regular']:.2f}")
        if not r["in_stock"]:
            notes.append("OOS")
        if r["variant"]:
            notes.append(r["variant"])
        rows.append({
            "compound": r["compound"], "vendor": r["vendor"], "method": r["method"],
            "mg": r["mg"], "mg_ml": None, "ml": None, "units": r["units"],
            "price": r["price"], "notes": " · ".join(notes) or None,
            "scraped": scraped,
            "per_mg": r["price"] / (r["mg"] * r["units"]),
        })

    wb = openpyxl.Workbook()
    n_v = len({r["vendor"] for r in rows})
    n_c = len({r["compound"] for r in rows})
    greens = write_tracker(
        sheet(wb, "Master Tracker"), rows,
        f"Last updated: {scraped:%Y-%m-%d %H:%M} UTC · {n_v} vendors · "
        f"{len(rows)} SKUs · scraped live from vendor APIs")

    kits = [r for r in rows if r["units"] >= 10]
    write_tracker(
        sheet(wb, "Full Kits", first=False), kits,
        f"FULL KITS (10+ units) · Last updated: {scraped:%Y-%m-%d %H:%M} UTC · "
        f"{len({r['vendor'] for r in kits})} vendors · {len(kits)} kit SKUs")

    # --- Price History: the thing the workbook alone cannot do.
    hs = sheet(wb, "Price History", first=False)
    hs["A1"] = "PRICE HISTORY — every sweep, append-only. This is what a spreadsheet can't hold."
    hs["A1"].font = Font(bold=True, size=11)
    for c, h in enumerate(["SCRAPED AT", "PEPTIDE", "VENDOR", "MG", "UNITS",
                           "PRICE", "REGULAR", "ON SALE", "IN STOCK"], 1):
        hs.cell(2, c, h).font = Font(bold=True, size=10)
        hs.column_dimensions[get_column_letter(c)].width = [20, 34, 24, 9, 9, 12, 12, 10, 11][c - 1]
    n = 0
    for i, r in enumerate(db.execute("""
            SELECT p.scraped_at, s.compound, s.vendor, s.mg, s.units,
                   p.price, p.regular, p.on_sale, p.in_stock
            FROM price_snapshots p JOIN skus s ON s.id = p.sku_id
            ORDER BY p.scraped_at DESC, s.compound LIMIT 20000""")):
        rw = 3 + i
        for c, v in enumerate([r["scraped_at"], r["compound"], r["vendor"], r["mg"],
                               r["units"], r["price"], r["regular"],
                               "yes" if r["on_sale"] else "", "yes" if r["in_stock"] else "no"], 1):
            cell = hs.cell(rw, c, v)
            cell.fill = PatternFill("solid", fgColor=BAND_A if i % 2 == 0 else BAND_B)
            if c in (6, 7):
                cell.number_format = '"$"#,##0.00'
        n += 1
    hs.auto_filter.ref = f"A2:I{2 + n}"
    hs.freeze_panes = "A3"

    # --- Changes
    cs = sheet(wb, "Changes", first=False)
    cs["A1"] = "PRICE CHANGES — detected by diffing consecutive sweeps. Drives the alerts."
    cs["A1"].font = Font(bold=True, size=11)
    for c, h in enumerate(["DETECTED", "PEPTIDE", "VENDOR", "WAS", "NOW", "% CHANGE", "KIND"], 1):
        cs.cell(2, c, h).font = Font(bold=True, size=10)
        cs.column_dimensions[get_column_letter(c)].width = [20, 34, 24, 12, 12, 12, 10][c - 1]
    m = 0
    for i, r in enumerate(db.execute("""
            SELECT c.detected_at, s.compound, s.vendor, c.old_price, c.new_price,
                   c.pct_delta, c.kind
            FROM price_changes c JOIN skus s ON s.id = c.sku_id
            ORDER BY c.detected_at DESC, c.pct_delta""")):
        rw = 3 + i
        for c, v in enumerate([r["detected_at"], r["compound"], r["vendor"],
                               r["old_price"], r["new_price"], r["pct_delta"] / 100,
                               r["kind"]], 1):
            cell = cs.cell(rw, c, v)
            if c in (4, 5):
                cell.number_format = '"$"#,##0.00'
            if c == 6:
                cell.number_format = "0.0%"
        m += 1
    if m == 0:
        cs["A3"] = ("No changes yet — this is the first sweep. Rows appear from the "
                    "second sweep onward, and this tab is what the email/SMS digest reads.")
        cs["A3"].font = Font(italic=True, color="FF666666")

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"wrote {args.out}")
    print(f"  Master Tracker : {len(rows)} SKUs · {n_v} vendors · {n_c} compounds · {greens} green")
    print(f"  Full Kits      : {len(kits)} rows")
    print(f"  Price History  : {n} snapshots")
    print(f"  Changes        : {m} rows")
    db.close()


if __name__ == "__main__":
    main()
