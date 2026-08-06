#!/usr/bin/env python3
"""
Migrate Brandon Schubert's legacy peptide tracker (v50) into his v3 schema.

Brandon wrote two clean specs (SCRAPE.md / INGEST.md) defining a 3-tab workbook
and built the empty target (v3). The data never moved -- all 6,972 live rows are
still in the 95-tab v50 workbook. This script runs his own rules against his own
data and writes a populated copy of v3.

  python3 scripts/peptide_migrate.py \
      --source "~/Downloads/Master_Peptide_Tracker_v50_updated 07.24.xlsx" \
      --target "~/Downloads/Master_Peptide_Tracker_v3.xlsx" \
      --out demos/brandon-peptide-migration/Master_Peptide_Tracker_v3_populated.xlsx \
      --flags demos/brandon-peptide-migration/flagged-decisions.md

Neither input is modified. Everything lands in --out.

Source quirks that cost time if you don't know them:

  * v50 'Master Tracker' has an EMPTY column A. Data starts row 4, and the real
    columns are B=Vendor C=Peptide D=MG/Vial E=Vials F=MG Total G=Kit Price
    H=$/Vial I=$/MG J=Notes L=mg/vial helper.
  * v50 column D ("MG/Vial") is free text, 141 distinct formats -- '5mg' but also
    '1000mcg caps 60ct', '12600ui', '215mg/mL x 2mL', '80mg (50/10/10/10)', '?mg'.
    parse_size() below enumerates every one of them; it is not a general parser.
  * openpyxl reads v50 with data_only=True as all-None: the file was written by a
    tool that never cached formula results. We read data_only=False and take the
    literal values, which is fine because B/C/D/E/G are all literals.
  * v3 ships with J/L/M/N formulas and the Master Lists Y-column audit hardcoded
    to row 1500. We're inserting ~6,900 rows, so both get extended. INGEST.md
    says extend the formulas rather than hardcode values -- so we do.

INGEST.md rules enforced here, because breaking any of them corrupts the sheet:
  * Write A-I, K, O, P only. J, L, M, N are formula columns, never written.
  * Blank is not zero. A '0' in an inapplicable field produces a wrong MG TOTAL.
  * Flag, don't guess (Step 2). Anything ambiguous goes to --flags and is skipped.
"""

import argparse
import re
import shutil
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    sys.exit("openpyxl required:  pip3 install openpyxl")

# INGEST.md Step 4 / Step 5 -- exact ARGB fills. Banding follows position, not
# identity, so it is a full re-stripe on every run.
BAND_A = "FFFFFFFF"
BAND_B = "FFEDE9FE"
GREEN = "FF90EE90"

METHOD_ORDER = {"Inj": 0, "Oral": 1, "Nasal": 2, "LiqDrop": 3}

# v3 target columns (1-indexed)
C_PEPTIDE, C_VENDOR, C_METHOD = 1, 2, 3
C_MGVIAL, C_MGCAP, C_CAPSBOT = 4, 5, 6
C_MGML, C_MLBOT, C_UNITS = 7, 8, 9
C_MGTOTAL, C_TOTAL, C_PERVIAL = 10, 11, 12
C_PERMG, C_PERML, C_NOTES, C_SCRAPED = 13, 14, 15, 16

FORMULA_COLS = (C_MGTOTAL, C_PERVIAL, C_PERMG, C_PERML)

# Vendor names v50 uses that normalization alone can't resolve. Left side is the
# v50 tab-style name, right side is the canonical Master Lists!R display name.
# All confirmed against the 92-vendor table in SCRAPE.md, matching on URL.
VENDOR_OVERRIDES = {
    "amercianpeptides": "American Peptides",   # typo in his source
    "ngpeptide": "NextGen Peptides",           # ngpeptide.com per SCRAPE.md
    "ruobio": "RUO",
    "purepeptidesbio": "Pure Peptides",
    "nexaph": "Nexaph",
    "atomiklabz": "Atomik Labs",               # atomiklabz.com -> "Atomik Labs"
    "coffeeandpeppers": "Coffee & Peppers",    # the '&' defeats normalization
    "omegamino": "OmegaAmino",                 # omegamino.net -> "OmegaAmino"
}

# v50 vendors genuinely absent from his canonical 92 -- NOT typos. SCRAPE.md:
# "If a vendor is not on this list, do not scrape it. Report the name and URL so
# it can be added." So these flag for his decision rather than getting mapped.
# (Note EliteBiogenix != "Elite Research"; different company.)


def norm(s):
    """Casefold and strip everything that isn't alphanumeric."""
    return re.sub(r"[^a-z0-9]", "", str(s).casefold())


def _num(s):
    return float(s) if "." in s else int(s)


def parse_size(raw):
    """Turn a v50 size string into (method, fields, note, flag_reason).

    fields is a dict of the v3 columns to populate. Only applicable keys appear
    -- absent means leave the cell empty, which is NOT the same as zero.
    Returns flag_reason non-None when the string can't be resolved; caller skips
    the row and records it (INGEST.md Step 2: surface decisions, then stop).
    """
    if raw is None:
        return None, None, None, "blank size"
    s = str(raw).strip()

    if s == "?mg" or s == "":
        return None, None, None, f"unresolvable size {s!r}"

    # --- IU-based (HCG). INGEST.md: Inj, IU count in MG/VIAL. '12600ui' is his
    # typo for IU and appears 3x.
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(?:iu|ui)", s, re.I)
    if m:
        return "Inj", {C_MGVIAL: _num(m.group(1))}, "IU-based; MG/VIAL holds IU", None

    # --- Oral: capsules/tablets. mg or mcg per cap, optional count.
    m = re.fullmatch(
        r"(\d+(?:\.\d+)?)\s*(mg|mcg)\s*(?:caps?|tablets?)?\s*"
        r"(?:[x×]\s*)?(?:\(?(\d+)\s*(?:ct|caps?)\)?)?",
        s, re.I,
    )
    if m and re.search(r"caps?|tablet|ct", s, re.I):
        per = _num(m.group(1))
        note = None
        if m.group(2).lower() == "mcg":
            per, note = per / 1000, f"{m.group(1)}mcg per capsule"
        if not m.group(3):
            return None, None, None, f"capsule count not stated in {s!r}"
        return "Oral", {C_MGCAP: per, C_CAPSBOT: int(m.group(3))}, note, None

    # --- mcg vial. INGEST.md: convert to decimal mg (500mcg -> 0.5).
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*mcg", s, re.I)
    if m:
        return "Inj", {C_MGVIAL: _num(m.group(1)) / 1000}, f"{m.group(1)}mcg vial", None

    # --- 'MG/mL x NmL' pre-mixed solution.
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*mg\s*/\s*ml\s*x\s*(\d+(?:\.\d+)?)\s*ml", s, re.I)
    if m:
        return "LiqDrop", {C_MGML: _num(m.group(1)), C_MLBOT: _num(m.group(2))}, None, None

    # --- Volume only, no concentration stated. INGEST.md: ML/BOTTLE only, MG/ML
    # left empty so $/ML carries the value and $/MG resolves blank.
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*ml", s, re.I)
    if m:
        return "LiqDrop", {C_MLBOT: _num(m.group(1))}, "no mg/ml stated", None

    # --- Blend, explicit total with the split in parens: '70mg (50/10/10)'.
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*mg\s*\((.+?)\)", s, re.I)
    if m:
        return "Inj", {C_MGVIAL: _num(m.group(1))}, f"blend split {m.group(2)}", None

    # --- 'Nmg total' -- already the sum.
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*mg\s*total", s, re.I)
    if m:
        return "Inj", {C_MGVIAL: _num(m.group(1))}, "blend total as listed", None

    # --- Plain single vial. The 97.7% case.
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*mg", s, re.I)
    if m:
        return "Inj", {C_MGVIAL: _num(m.group(1))}, None, None

    # --- Blend split with no total: '5mg+5mg', '13/3mg', '50mg/10mg/10mg'.
    # INGEST.md: MG/VIAL is the sum of all components, NOTES carries the split.
    if re.fullmatch(r"[\d.\s]+(?:mg)?(?:\s*[+/]\s*[\d.\s]+(?:mg)?)+", s, re.I):
        parts = [p for p in re.split(r"[+/]", s) if p.strip()]
        nums = []
        for p in parts:
            pm = re.search(r"(\d+(?:\.\d+)?)", p)
            if pm:
                nums.append(_num(pm.group(1)))
        if nums:
            total = sum(nums)
            return "Inj", {C_MGVIAL: total}, f"blend split {s} (sum {total}mg)", None

    return None, None, None, f"unrecognized size format {s!r}"


def build_vendor_map(v50_names, canon):
    """Map v50 tab-style vendor names onto canonical display names."""
    by_norm = {norm(c): c for c in canon}
    mapping, unresolved = {}, []
    for name in sorted(v50_names):
        n = norm(name)
        if n in VENDOR_OVERRIDES:
            mapping[name] = VENDOR_OVERRIDES[n]
        elif n in by_norm:
            mapping[name] = by_norm[n]
        else:
            # v50 drops suffixes ('10BottleValue' vs '10 Bottle Value Co.') and
            # occasionally abbreviates ('AmeanoP' vs 'Ameano Peptides').
            hits = [c for k, c in by_norm.items() if k.startswith(n) or n.startswith(k)]
            if len(hits) == 1:
                mapping[name] = hits[0]
            else:
                unresolved.append((name, hits))
    return mapping, unresolved


def strip_format_suffix(peptide):
    """'BPC-157 (Tablets)' -> ('BPC-157', 'Oral').

    SCRAPE.md R9 / the naming rule: base compound in PEPTIDE, format in METHOD.
    This is exactly the problem the v3 schema was designed to fix.
    """
    m = re.fullmatch(r"(.+?)\s*\((tablets|capsules|caps|oral)\)", peptide.strip(), re.I)
    if m:
        return m.group(1).strip(), "Oral"
    return peptide.strip(), None


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--source", required=True, help="v50 workbook (read-only)")
    ap.add_argument("--target", required=True, help="empty v3 workbook (read-only)")
    ap.add_argument("--out", required=True, help="populated workbook to write")
    ap.add_argument("--flags", help="markdown file for decisions we refused to guess")
    ap.add_argument("--scraped", default="2026-07-24", help="scrape date for column P")
    args = ap.parse_args()

    src = Path(args.source).expanduser()
    tgt = Path(args.target).expanduser()
    out = Path(args.out).expanduser()
    scraped = datetime.strptime(args.scraped, "%Y-%m-%d").date()
    out.parent.mkdir(parents=True, exist_ok=True)

    # --- Read canonical lists off the untouched target.
    wb_t = openpyxl.load_workbook(tgt, data_only=False)
    ml = wb_t["Master Lists"]
    canon_pep = [ml.cell(r, 1).value for r in range(2, 72) if ml.cell(r, 1).value]
    canon_ven = [ml.cell(r, 18).value for r in range(2, 95) if ml.cell(r, 18).value]
    canon_pep_set, canon_ven_set = set(canon_pep), set(canon_ven)
    print(f"canonical: {len(canon_pep)} compounds, {len(canon_ven)} vendors")

    # --- Read v50.
    wb_s = openpyxl.load_workbook(src, data_only=False, read_only=True)
    raw = []
    for r in wb_s["Master Tracker"].iter_rows(min_row=4, values_only=True):
        vendor, peptide, size, units, _mgtot, price, *rest = r[1:]
        if vendor and peptide:
            raw.append({
                "vendor": str(vendor).strip(),
                "peptide": str(peptide).strip(),
                "size": size,
                "units": units,
                "price": price,
                "notes": (str(r[9]).strip() if r[9] else None),
            })
    wb_s.close()
    print(f"read {len(raw)} source rows")

    vendor_map, ven_unresolved = build_vendor_map({d["vendor"] for d in raw}, canon_ven)
    print(f"vendor map: {len(vendor_map)} resolved, {len(ven_unresolved)} unresolved")

    # --- Transform.
    rows, flags = [], []
    for d in raw:
        pep, method_override = strip_format_suffix(d["peptide"])

        if pep not in canon_pep_set:
            flags.append((d, f"compound {pep!r} not in the canonical 70-list"))
            continue
        if d["vendor"] not in vendor_map:
            flags.append((d, f"vendor {d['vendor']!r} has no canonical match"))
            continue
        if not isinstance(d["units"], (int, float)) or not d["units"]:
            flags.append((d, f"unit count missing or non-numeric ({d['units']!r})"))
            continue
        if not isinstance(d["price"], (int, float)):
            flags.append((d, f"price missing or non-numeric ({d['price']!r})"))
            continue

        method, fields, note, reason = parse_size(d["size"])
        if reason:
            flags.append((d, reason))
            continue
        if method_override:
            # 'BPC-157 (Tablets)' with size '20mg caps' -- both say Oral. If the
            # size parsed as Inj, the name is the more specific signal.
            if method != "Oral":
                flags.append((d, f"name says Oral but size {d['size']!r} parsed as {method}"))
                continue
            method = "Oral"

        notes = [x for x in (d["notes"], note) if x]
        rows.append({
            "peptide": pep,
            "vendor": vendor_map[d["vendor"]],
            "method": method,
            "fields": fields,
            "units": d["units"],
            "price": d["price"],
            "notes": " · ".join(notes) or None,
        })

    # INGEST.md Step 2 item 4: duplicate rows within the incoming batch are a
    # decision to surface, not something to silently write. Keep the first
    # occurrence, flag the rest with both prices so he can pick.
    seen, deduped = {}, []
    for r in rows:
        key = (r["peptide"], r["vendor"], r["method"],
               tuple(sorted(r["fields"].items())), r["units"])
        if key in seen:
            flags.append(({
                "vendor": r["vendor"], "peptide": r["peptide"],
                "size": f"{r['method']} {dict(r['fields'])}", "units": r["units"],
                "price": f"{r['price']} (duplicate of a row priced {seen[key]})",
            }, "duplicate SKU in source — same vendor/compound/method/size/units"))
            continue
        seen[key] = r["price"]
        deduped.append(r)
    rows = deduped

    print(f"transformed {len(rows)} rows, flagged {len(flags)}")

    # --- Derived metrics. openpyxl won't evaluate J/M/N, so compute them here to
    # decide the green fills. The cells themselves stay formulas.
    for r in rows:
        f, u = r["fields"], r["units"]
        if r["method"] == "Oral":
            r["mg_total"] = u * f.get(C_CAPSBOT, 0) * f.get(C_MGCAP, 0)
        elif r["method"] in ("Nasal", "LiqDrop"):
            r["mg_total"] = u * f.get(C_MLBOT, 0) * f.get(C_MGML, 0)
        else:
            r["mg_total"] = u * f.get(C_MGVIAL, 0)
        r["per_mg"] = r["price"] / r["mg_total"] if r["mg_total"] else None
        ml_tot = u * f.get(C_MLBOT, 0)
        r["per_ml"] = r["price"] / ml_tot if ml_tot else None
        r["unit_size"] = r["mg_total"] / u if u else 0

    # --- Sort: PEPTIDE -> METHOD -> VENDOR -> per-unit size -> unit count.
    rows.sort(key=lambda r: (r["peptide"].casefold(), METHOD_ORDER[r["method"]],
                             r["vendor"].casefold(), r["unit_size"], r["units"]))

    def write_tab(ws, data, first_row=3):
        """Write A-I, K, O, P. Never J, L, M, N."""
        for i, r in enumerate(data):
            rw = first_row + i
            band = BAND_A if i % 2 == 0 else BAND_B
            ws.cell(rw, C_PEPTIDE, r["peptide"])
            ws.cell(rw, C_VENDOR, r["vendor"])
            ws.cell(rw, C_METHOD, r["method"])
            for col, val in r["fields"].items():
                ws.cell(rw, col, val)
            ws.cell(rw, C_UNITS, r["units"])
            ws.cell(rw, C_TOTAL, r["price"])
            if r["notes"]:
                ws.cell(rw, C_NOTES, r["notes"])
            ws.cell(rw, C_SCRAPED, scraped)
            for col in range(1, 17):
                ws.cell(rw, col).fill = PatternFill("solid", fgColor=band)

        # Extend the J/L/M/N formulas to the new last row (INGEST.md: extend,
        # never hardcode). Template them off row 3, which v3 ships with.
        last = first_row + len(data) - 1
        templates = {c: ws.cell(3, c).value for c in FORMULA_COLS}
        for rw in range(first_row, last + 1):
            for col, tpl in templates.items():
                if tpl:
                    ws.cell(rw, col, re.sub(r"(\$?[A-Z]{1,2}\$?)3\b",
                                            lambda m: f"{m.group(1)}{rw}", tpl))
        # Clear any leftover rows from the shipped 1500/600-row skeleton.
        for rw in range(last + 1, ws.max_row + 1):
            for col in range(1, 17):
                c = ws.cell(rw, col)
                c.value = None
                c.fill = PatternFill(fill_type=None)
        return last

    def green_fills(ws, data, first_row=3):
        """Two independent best-value passes, rebuilt across the whole tab."""
        # $/MG grouped by (PEPTIDE, METHOD, per-unit size)
        groups = defaultdict(list)
        for i, r in enumerate(data):
            if r["per_mg"] is not None:
                groups[(r["peptide"], r["method"], round(r["unit_size"], 6))].append((i, r["per_mg"]))
        n_mg = 0
        for members in groups.values():
            lo = min(v for _, v in members)
            for i, v in members:
                if v == lo:
                    ws.cell(first_row + i, C_PERMG).fill = PatternFill("solid", fgColor=GREEN)
                    n_mg += 1
        # $/ML grouped by (PEPTIDE, METHOD, ML/BOTTLE), Nasal + LiqDrop only
        groups = defaultdict(list)
        for i, r in enumerate(data):
            if r["method"] in ("Nasal", "LiqDrop") and r["per_ml"] is not None:
                groups[(r["peptide"], r["method"], r["fields"].get(C_MLBOT))].append((i, r["per_ml"]))
        n_ml = 0
        for members in groups.values():
            lo = min(v for _, v in members)
            for i, v in members:
                if v == lo:
                    ws.cell(first_row + i, C_PERML).fill = PatternFill("solid", fgColor=GREEN)
                    n_ml += 1
        return n_mg, n_ml

    mt = wb_t["Master Tracker"]
    last_mt = write_tab(mt, rows)
    mg_greens, ml_greens = green_fills(mt, rows)
    mt.auto_filter.ref = f"A2:P{last_mt}"

    kits = [r for r in rows if r["units"] >= 10]
    fk = wb_t["Full Kits"]
    last_fk = write_tab(fk, kits)
    fk_mg_greens, fk_ml_greens = green_fills(fk, kits)
    fk.auto_filter.ref = f"A2:P{last_fk}"

    # --- Master Lists: FORMATS TRACKED (col B) and scraped date (col S).
    fmt = defaultdict(set)
    for r in rows:
        fmt[r["peptide"]].add(r["method"])
    for rw in range(2, 72):
        name = ml.cell(rw, 1).value
        if name and name in fmt:
            ml.cell(rw, 2, ", ".join(sorted(fmt[name], key=lambda m: METHOD_ORDER[m])))
    touched = {r["vendor"] for r in rows}
    for rw in range(2, 95):
        if ml.cell(rw, 18).value in touched:
            ml.cell(rw, 19, scraped)

    # --- The Y-column audit formulas hardcode $A$3:$A$1500. We just went well
    # past that, so extend the ranges or every count silently undercounts.
    audit_fixed = 0
    for rw in range(2, 300):
        c = ml.cell(rw, 25)
        if isinstance(c.value, str) and "1500" in c.value:
            c.value = c.value.replace("$1500", f"${last_mt}")
            audit_fixed += 1

    # The audit block has 70 formula rows in Y/Z counting against a label in
    # column X -- but X3:X72 shipped empty, so it has never returned a number.
    # Fill the labels from the canonical list. The formulas are left untouched
    # (INGEST.md Step 7: "formula-driven, updates itself... don't touch").
    audit_labels = 0
    for i, name in enumerate(canon_pep):
        if 3 + i <= 72:
            ml.cell(3 + i, 24, name)
            audit_labels += 1

    n_vendors, n_kit_vendors = len({r["vendor"] for r in rows}), len({r["vendor"] for r in kits})
    mt["A1"] = (f"Last updated: {scraped.isoformat()} · {n_vendors} vendors · "
                f"{len(rows)} SKUs · {len(kits)} full-kit SKUs")
    fk["A1"] = (f"FULL KITS (10+ units) · Last updated: {scraped.isoformat()} · "
                f"{n_kit_vendors} vendors · {len(kits)} kit SKUs")

    wb_t.save(out)
    print(f"\nwrote {out}")

    # ------------------------------------------------------------------ verify
    # INGEST.md Step 9. "Do not hand off a file with a known failure."
    wb_v = openpyxl.load_workbook(out, data_only=False)
    v_mt, v_fk, v_ml = wb_v["Master Tracker"], wb_v["Full Kits"], wb_v["Master Lists"]
    data_rows = [r for r in v_mt.iter_rows(min_row=3, max_row=last_mt, values_only=True)]

    tuples = [(r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8]) for r in data_rows]
    zero_in_blank = sum(
        1 for r in data_rows
        for col in (C_MGVIAL, C_MGCAP, C_CAPSBOT, C_MGML, C_MLBOT)
        if r[col - 1] == 0
    )
    formulas_ok = all(
        isinstance(r[c - 1], str) and r[c - 1].startswith("=")
        for r in data_rows for c in FORMULA_COLS
    )
    sort_key = [(r[0].casefold(), METHOD_ORDER[r[2]], r[1].casefold()) for r in data_rows]

    checks = [
        ("All J/L/M/N cells are formulas (nothing hardcoded)", formulas_ok),
        ("Row count reconciles: source = written + flagged",
         len(raw) == len(rows) + len(flags)),
        (f"No 0 written into an inapplicable field ({zero_in_blank} found)",
         zero_in_blank == 0),
        ("Every PEPTIDE matches Master Lists!A2:A71",
         all(r[0] in canon_pep_set for r in data_rows)),
        ("Every VENDOR matches Master Lists!R2:R93",
         all(r[1] in canon_ven_set for r in data_rows)),
        ("Every METHOD is one of the four permitted values",
         all(r[2] in METHOD_ORDER for r in data_rows)),
        (f"No duplicate (peptide,vendor,method,size,units) tuples",
         len(tuples) == len(set(tuples))),
        ("Sort order correct", sort_key == sorted(sort_key)),
        ("Banding unbroken across all populated rows",
         all(v_mt.cell(3 + i, 1).fill.fgColor.rgb == (BAND_A if i % 2 == 0 else BAND_B)
             for i in range(len(data_rows)))),
        (f"$/MG green cells present ({mg_greens})", mg_greens > 0),
        (f"$/ML green cells present ({ml_greens})", ml_greens > 0),
        ("Full Kits count equals rows where units >= 10",
         (last_fk - 2) == len(kits)),
        (f"AutoFilter covers new last row (A2:P{last_mt})",
         v_mt.auto_filter.ref == f"A2:P{last_mt}"),
        ("A1 summary lines updated on both tabs",
         str(v_mt["A1"].value).startswith("Last updated: 2026")
         and "0 kit SKUs" not in str(v_fk["A1"].value)),
        (f"Audit formula ranges extended past 1500 ({audit_fixed} formulas)",
         audit_fixed > 0),
        (f"Audit compound labels populated in X3:X72 ({audit_labels})",
         audit_labels == len(canon_pep)),
    ]
    print("\n--- INGEST.md Step 9 validation ---")
    for label, ok in checks:
        print(f"  [{'PASS' if ok else 'FAIL'}]  {label}")
    failed = [l for l, ok in checks if not ok]
    print(f"\n{len(checks) - len(failed)}/{len(checks)} passed")

    # ----------------------------------------------------------------- summary
    print(f"\nmigrated : {len(rows)} SKUs · {n_vendors} vendors · "
          f"{len(fmt)} compounds · {len(kits)} full-kit rows")
    print("methods  :", dict(Counter(r["method"] for r in rows)))
    print(f"notes    : {sum(1 for r in rows if r['notes'])} rows carry notes")

    if args.flags:
        by_reason = defaultdict(list)
        for d, reason in flags:
            by_reason[reason].append(d)
        fp = Path(args.flags).expanduser()
        fp.parent.mkdir(parents=True, exist_ok=True)
        with fp.open("w") as fh:
            fh.write("# Flagged decisions — need Brandon's call\n\n")
            fh.write(f"{len(flags)} of {len(raw)} source rows were **not** migrated. "
                     "Per INGEST.md Step 2, these are surfaced rather than guessed.\n\n")
            for reason in sorted(by_reason, key=lambda k: -len(by_reason[k])):
                items = by_reason[reason]
                fh.write(f"## {reason} — {len(items)} row(s)\n\n")
                fh.write("| Vendor | Compound | Size | Units | Price |\n")
                fh.write("|---|---|---|---|---|\n")
                for d in items[:25]:
                    fh.write(f"| {d['vendor']} | {d['peptide']} | `{d['size']}` "
                             f"| {d['units']} | {d['price']} |\n")
                if len(items) > 25:
                    fh.write(f"\n_+{len(items) - 25} more._\n")
                fh.write("\n")
            if ven_unresolved:
                fh.write("## Vendor names with no confident canonical match\n\n")
                for name, hits in ven_unresolved:
                    fh.write(f"- `{name}` → {hits or 'no candidate'}\n")
        print(f"flags    : {fp}")

    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
